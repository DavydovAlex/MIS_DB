import abc
import sys

from openpyxl import Workbook, \
    load_workbook
from abc import ABC


class FileHandler(ABC):
    _path = None
    _header = None

    def __init__(self, path):
        self._path = path

    @property
    def header(self):
        return self._header

    @property
    def path(self):
        return self._path

    def set_header(self, val):
        self._header = val
        return self

    @abc.abstractmethod
    def write(self, iterator):
        pass


class ExcelHandler(FileHandler):
    MAX_ROWS = 999999  # Максимальное количество строк на листе
    _sheet_rows = None  # Количество строк на листе
    _wb = None

    def __init__(self, path):
        super().__init__(path)
        self._sheet_rows = self.MAX_ROWS

    @property
    def sheet_rows(self):
        return self._sheet_rows

    def set_sheet_rows(self, val):
        if type(val) != int:
            raise TypeError('Rows number must be int')
        if val >= 1000000:
            raise ValueError('Maximum rows count is 999999')
        self._sheet_rows = val

    def write(self, iterator):
        self.write_data(iterator, iterator.header)

    def write_data(self, data, header=None):
        self._wb = Workbook(write_only=True)
        ws = self._wb.create_sheet()
        rows_in_sheet = 0
        if header:
            ws.append(header)
            rows_in_sheet = 1
        for row in data:
            if rows_in_sheet % 100000 == 0:
                print(str(rows_in_sheet) + ':' + str(row))
            if rows_in_sheet < self._sheet_rows:
                rows_in_sheet += 1
                ws.append(row)
            else:
                rows_in_sheet = 0
                ws = self._wb.create_sheet()
                if header:
                    ws.append(header)
                    rows_in_sheet = 1
                rows_in_sheet += 1
                ws.append(row)
        self._wb.save(self.path)


    def read(self):
        self._wb = load_workbook(self.path)

    def get_raw_page(self, page_number=0):
        sheet = self._wb.worksheets[page_number]
        return sheet.iter_rows(values_only=True)

    def handle_page(self, page_number=0):
        sheet = self._wb.worksheets[page_number]
        self._remove_empty_columns(sheet)
        self._remove_empty_rows(sheet)
        return sheet.iter_rows(values_only=True)

    def _remove_empty_rows(self, sheet):
        rows_to_remove = []
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            column_values = set(row)

            if len(column_values) == 1 and None in column_values:
                print(str(i) + str(column_values))
                rows_to_remove.append(i)
        for row_number in sorted(rows_to_remove, reverse=True):
            sheet.delete_rows(row_number)

    def _remove_empty_columns(self, sheet):
        columns_to_remove = []
        for i, col in enumerate(sheet.iter_cols(values_only=True), start=1):
            column_values = set(col)
            print(column_values)
            if len(column_values) == 1 and None in column_values:
                columns_to_remove.append(i)
        for column_number in sorted(columns_to_remove, reverse=True):
            sheet.delete_cols(column_number)

    def get_data(self, sheet, data_start_row):
        return sheet.iter_rows(min_row=data_start_row, values_only=True)

    def get_header(self, header_row):
        if header_row:
            wb = load_workbook(self.path)
            sheet = wb.worksheets[0]
            max_row = sheet.max_row
            max_column = sheet.max_column


if __name__ == '__main__':
    ex = ExcelHandler(r'D:\tmp\Children01.07.22_1656993384.xlsx')
    ex.read()
    for row in ex.handle_page():
        print(row)
